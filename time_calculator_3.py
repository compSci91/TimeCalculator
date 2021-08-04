import openpyxl

wb = openpyxl.load_workbook("Trelle Dandridge.xlsx")
sheet = wb["Sheet1"]

max_number_of_rows = sheet.max_row

row = 3
max_number_of_rows = sheet.max_row
total_touch_time = 0
print(max_number_of_rows)

while row <= max_number_of_rows:
    speech_string = sheet.cell(row,5).value
    if(speech_string):
        start_row = row
        potential_end_row = row + 1

        should_continue_searching_for_end_row = True

        while should_continue_searching_for_end_row:
            potential_end_row_is_touching_element = sheet.cell(potential_end_row, 5).value

            if(potential_end_row_is_touching_element):
                potential_end_row = potential_end_row + 1

            else:
                end_row = potential_end_row - 1
                start_time = float(sheet.cell(start_row, 3).value)
                end_time = float(sheet.cell(end_row, 3).value)

                # print(f'Start row {start_row} -- End row {end_row}')
                print(end_time - start_time)
                total_touch_time = total_touch_time + (end_time - start_time)
                should_continue_searching_for_end_row = False
                row = potential_end_row
    else:
        row = row + 1


print(total_touch_time)
total_time = sheet.cell(max_number_of_rows, 3).value - sheet.cell(3,3).value
print(total_time)

print(f'Total element touch time {total_touch_time}')
print(f'Total time {total_time}')
print(f'Percentage of time: {total_touch_time / total_time}')
print(f'Percentage of time NOT touching: { 1 - total_touch_time / total_time}')
