from datetime import datetime, date, time, timedelta
import re
import openpyxl


def convert_time_string(time_string):
    time_array = re.split('[,:]', time_string)
    return time(int(time_array[0]), int(time_array[1]), int(time_array[2]), int(time_array[3])*100000)

def calculate_time_difference(start_time_string, end_time_string):
    start_time = convert_time_string(start_time_string)
    end_time = convert_time_string(end_time_string)
    return datetime.combine(date.min, end_time) - datetime.combine(date.min, start_time)

def convert_duration_string(duration_string):
    time_array = re.split('[,:]', duration_string)
    # return timedelta(int(time_array[0]), int(time_array[1]), int(time_array[2]), int(time_array[3]))

    microseconds =  int(time_array[3]) if len(time_array) == 4 else 0

    return timedelta(weeks=0, days=0, hours=int(time_array[0]), minutes=int(time_array[1]), seconds=int(time_array[2]), milliseconds=0, microseconds = microseconds)


        # return time(int(time_array[0]), int(time_array[1]), int(time_array[2]), int(time_array[3]))



# print(calculate_time_difference("0:00:09,4", "0:00:10,7"))
wb = openpyxl.load_workbook("Amanda.xlsx")

sheet = wb["Sheet1"]

max_number_of_rows = sheet.max_row

#DO NOT ERASE This calculates duration per row
# for row in range(2, max_number_of_rows + 1):
#     start_time_string = sheet.cell(row,2).value
#     end_time_string = sheet.cell(row,3).value
#     time_difference = str(calculate_time_difference(start_time_string, end_time_string)).replace(".", ",", 1)
#     sheet.cell(row, 4, value=time_difference)




for expression_number in range(1,15):
    total_duration = timedelta(weeks=0, days=0, hours=0, minutes=0, seconds=0, milliseconds=0, microseconds=0)
    for row in range(2, max_number_of_rows + 1):
        if(sheet.cell(row, 1).value == f"{expression_number}. Touching expression" or sheet.cell(row, 1).value == f"{expression_number}.Touching expression"):
            current_duration_string = sheet.cell(row,4).value
            # print(current_duration_string)
            current_duration = convert_duration_string(current_duration_string)
            total_duration = total_duration + current_duration

    print(f"{expression_number}. Touching expression: ", total_duration)

    total_duration = timedelta(weeks=0, days=0, hours=0, minutes=0, seconds=0, milliseconds=0, microseconds=0)
    for row in range(2, max_number_of_rows + 1):
        if(sheet.cell(row, 1).value == f"{expression_number}. Not touching expression" or sheet.cell(row, 1).value == f"{expression_number}.Not touching expression"):
            current_duration_string = sheet.cell(row,4).value
            current_duration = convert_duration_string(current_duration_string)
            total_duration = total_duration + current_duration

    print(f"{expression_number}. Not touching expression: ", total_duration)
    print("")




wb.save("Amanda.xlsx")
