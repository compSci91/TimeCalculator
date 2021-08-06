import openpyxl
import sys

file_name = sys.argv[1]
print(file_name)
wb = openpyxl.load_workbook(file_name)

EXPRESSION_NUMBER = 1
TOTAL_EXPRESSION_TIME = 2
TOTAL_TIME_TOUCHING_EXPRESSION = 3
PERCENTAGE_OF_TIME_TOUCHING_EXPRESSION = 4
PERCENTAGE_OF_TIME_NOT_TOUCHING_EXPRESSION = 5
TOTAL_ELEMENT_TOUCH_TIME = 6
PERCENTAGE_OF_TIME_TOUCHING_ELEMENT = 7
PERCENTAGE_OF_TIME_NOT_TOUCHING_ELEMENT = 8
PERCENTAGE_OF_TIME_TOUCHING_ELEMENT_GIVEN_TOUCHING_EXPRESSION = 9
PERCENTAGE_OF_TIME_NOT_TOUCHING_ELEMENT_GIVEN_TOUCHING_EXPRESSION = 10

info_sheet = wb["Sheet15"]
info_sheet.cell(row=1, column=EXPRESSION_NUMBER, value="EXPRESSION NUMBER")
info_sheet.cell(row=1, column=TOTAL_EXPRESSION_TIME, value="TOTAL EXPRESSION TIME")
info_sheet.cell(row=1, column=TOTAL_TIME_TOUCHING_EXPRESSION, value="TOTAL TIME TOUCHING EXPRESSION")
info_sheet.cell(row=1, column=PERCENTAGE_OF_TIME_TOUCHING_EXPRESSION, value="PERCENTAGE OF TIME TOUCHING EXPRESSION")
info_sheet.cell(row=1, column=PERCENTAGE_OF_TIME_NOT_TOUCHING_EXPRESSION, value="PERCENTAGE OF TIME NOT TOUCHING EXPRESSION")
info_sheet.cell(row=1, column=TOTAL_ELEMENT_TOUCH_TIME, value="TOTAL ELEMENT TOUCH TIME")
info_sheet.cell(row=1, column=PERCENTAGE_OF_TIME_TOUCHING_ELEMENT, value="PERCENTAGE OF TIME TOUCHING ELEMENT")
info_sheet.cell(row=1, column=PERCENTAGE_OF_TIME_NOT_TOUCHING_ELEMENT, value="PERCENTAGE OF TIME NOT TOUCHING ELEMENT")
info_sheet.cell(row=1, column=PERCENTAGE_OF_TIME_TOUCHING_ELEMENT_GIVEN_TOUCHING_EXPRESSION, value="PERCENTAGE OF TIME TOUCHING ELEMENT GIVEN TOUCHING EXPRESSION")
info_sheet.cell(row=1, column=PERCENTAGE_OF_TIME_NOT_TOUCHING_ELEMENT_GIVEN_TOUCHING_EXPRESSION, value="PERCENTAGE OF TIME NOT TOUCHING AN ELEMENT GIVEN TOUCHING EXPRESSION")



for sheet_number in range(1, 15):
    sheet_name = "Sheet" + str(sheet_number)
    sheet = wb[sheet_name]

    max_number_of_rows = sheet.max_row

    row = 3
    max_number_of_rows = sheet.max_row
    total_touch_time = 0

    while row <= max_number_of_rows:
        is_touching_expression = sheet.cell(row,4).value
        if(is_touching_expression):
            start_row = row
            potential_end_row = row + 1

            should_continue_searching_for_end_row = True

            while should_continue_searching_for_end_row:
                potential_end_row_is_touching_expression = sheet.cell(potential_end_row, 4).value

                if(potential_end_row_is_touching_expression):
                    potential_end_row = potential_end_row + 1

                else:
                    end_row = potential_end_row - 1
                    print(f'Start row: {start_row}')
                    start_time = float(sheet.cell(start_row, 3).value)
                    end_time = float(sheet.cell(end_row, 3).value)

                    # print(f'Start row {start_row} -- End row {end_row}')
                    # print(end_time - start_time)
                    total_touch_time = total_touch_time + (end_time - start_time)
                    should_continue_searching_for_end_row = False
                    row = potential_end_row
        else:
            row = row + 1



    print("***********************************")
    print(sheet_name)
    print(f'Number of rows: {max_number_of_rows}')

    total_time = sheet.cell(max_number_of_rows, 3).value - sheet.cell(3,3).value
    percentage_of_time_touching_expression = total_touch_time / total_time
    percentage_of_time_not_touching_expression = 1 - percentage_of_time_touching_expression

    print(f'Total touch time {total_touch_time}')
    print(f'Total Expression time {total_time}')
    print(f'Percentage of time touching expression: {percentage_of_time_touching_expression}')
    print(f'Percentage of time NOT touching expression: {percentage_of_time_not_touching_expression}')
    print("***********************************")

    expression_row_number = sheet_number + 1
    info_sheet.cell(row= expression_row_number, column=EXPRESSION_NUMBER, value= sheet_number)
    info_sheet.cell(row= expression_row_number, column=TOTAL_EXPRESSION_TIME, value= total_time)
    info_sheet.cell(row= expression_row_number, column=TOTAL_TIME_TOUCHING_EXPRESSION, value= total_touch_time)
    info_sheet.cell(row= expression_row_number, column=PERCENTAGE_OF_TIME_TOUCHING_EXPRESSION, value=percentage_of_time_touching_expression)
    info_sheet.cell(row= expression_row_number, column=PERCENTAGE_OF_TIME_NOT_TOUCHING_EXPRESSION, value=percentage_of_time_not_touching_expression)


    # sheet.cell(row=7, column=1, value="chess")
wb.save(file_name)


for sheet_number in range(1, 15):
    sheet_name = "Sheet" + str(sheet_number)
    sheet = wb[sheet_name]

    max_number_of_rows = sheet.max_row

    row = 3
    # max_number_of_rows = sheet.max_row
    total_element_touch_time = 0
    # print(max_number_of_rows)

    while row <= max_number_of_rows:
        speech_string = sheet.cell(row,5).value
        if(speech_string):
            start_row = row
            potential_end_row = row + 1

            should_continue_searching_for_end_row = True

            while should_continue_searching_for_end_row:
                potential_end_row_is_touching_element = sheet.cell(potential_end_row, 5).value

                if(potential_end_row_is_touching_element):
                    potential_end_row = potential_end_row + 1

                else:
                    end_row = potential_end_row - 1
                    start_time = float(sheet.cell(start_row, 3).value)
                    end_time = float(sheet.cell(end_row, 3).value)

                    # print(f'Start row {start_row} -- End row {end_row}')
                    # print(end_time - start_time)
                    total_element_touch_time = total_element_touch_time + (end_time - start_time)
                    should_continue_searching_for_end_row = False
                    row = potential_end_row
        else:
            row = row + 1


    print("***********************************")
    print(sheet_name)

    should_continue_searching_for_max_rows = True

    while(should_continue_searching_for_max_rows):
        if(sheet.cell(max_number_of_rows, 3).value):
            print("Max Number of Rows calculated CORRECTLY")
            should_continue_searching_for_max_rows = False

        else:
            print("max Number of rows calculated INCORRECTLY")
            max_number_of_rows = max_number_of_rows - 1

    print(f'Number of rows: {max_number_of_rows}')

    total_time = sheet.cell(max_number_of_rows, 3).value - sheet.cell(3,3).value
    percentage_of_time_touching_an_element = total_element_touch_time / total_time
    percentage_of_time_not_touching_an_elemnt = 1 - percentage_of_time_touching_an_element

    print(f'Total element touch time {total_element_touch_time}')
    print(f'Total Expression time {total_time}')
    print(f'Percentage of time touching an element: {percentage_of_time_touching_an_element}')
    print(f'Percentage of time NOT touching an element: {percentage_of_time_not_touching_an_elemnt}')
    print("***********************************")

    expression_row_number = sheet_number + 1
    total_time_touching_expression = info_sheet.cell(expression_row_number, TOTAL_TIME_TOUCHING_EXPRESSION).value

    percentage_of_time_touching_an_element_given_touching_an_expression = total_element_touch_time / total_time_touching_expression
    percentage_of_time_not_touching_an_element_given_touching_an_expression = 1 - percentage_of_time_touching_an_element_given_touching_an_expression

    info_sheet.cell(row= expression_row_number, column=TOTAL_ELEMENT_TOUCH_TIME, value= total_element_touch_time)
    info_sheet.cell(row= expression_row_number, column=PERCENTAGE_OF_TIME_TOUCHING_ELEMENT, value=percentage_of_time_touching_an_element)
    info_sheet.cell(row= expression_row_number, column=PERCENTAGE_OF_TIME_NOT_TOUCHING_ELEMENT, value=percentage_of_time_not_touching_an_elemnt)

    info_sheet.cell(row= expression_row_number, column=PERCENTAGE_OF_TIME_TOUCHING_ELEMENT_GIVEN_TOUCHING_EXPRESSION, value=percentage_of_time_touching_an_element_given_touching_an_expression)
    info_sheet.cell(row= expression_row_number, column=PERCENTAGE_OF_TIME_NOT_TOUCHING_ELEMENT_GIVEN_TOUCHING_EXPRESSION, value=percentage_of_time_not_touching_an_element_given_touching_an_expression)

wb.save(file_name)
