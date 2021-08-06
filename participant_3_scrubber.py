import openpyxl
import sys

file_name =  "./Excel Files/Participant 3 Paths.xlsx"
print(file_name)
wb = openpyxl.load_workbook(file_name)

sheet = wb["Sheet2"]

last_row = 325
last_time = float(sheet.cell(last_row, 3).value)

max_number_of_rows = sheet.max_row

print(max_number_of_rows)

for row_number in range(last_row + 1, max_number_of_rows):
    current_time = float(sheet.cell(row_number,3).value)
    new_time = last_time + current_time
    sheet.cell(row=row_number, column=3, value=new_time)

wb.save(file_name)
